"""Lightweight text extraction for metadata preview.

Extracts the first ~4000 characters from a file using local libraries only.
This is intentionally separate from the full ExtractionOrchestrator (Phase 2).
No UUID tracking, no structural metadata — just raw text for LLM consumption.
"""

import io
import re
import structlog

logger = structlog.get_logger()

# Maximum characters to extract for preview
MAX_PREVIEW_CHARS = 4000


def extract_preview(file_bytes: bytes, file_key: str) -> str:
    """Extract first 4000 characters of text from file bytes.

    Routes to format-specific extractor based on file extension.

    Args:
        file_bytes: Raw file content downloaded from MinIO.
        file_key: MinIO object key (used to determine file extension).

    Returns:
        Raw text string capped at 4000 characters.

    Raises:
        ValueError: If file format is not supported.
    """
    ext = _get_extension(file_key)

    extractor_map = {
        "md": _extract_markdown,
        "txt": _extract_text,
        "pdf": _extract_pdf,
        "docx": _extract_docx,
        "xlsx": _extract_xlsx,
        "png": _extract_image_ocr,
        "jpg": _extract_image_ocr,
        "jpeg": _extract_image_ocr,
    }

    extractor = extractor_map.get(ext)
    if not extractor:
        raise ValueError(f"Unsupported file format for preview: .{ext}")

    logger.info("preview_extraction_started", file_key=file_key, format=ext)
    text = extractor(file_bytes)
    preview = text[:MAX_PREVIEW_CHARS].strip()
    logger.info(
        "preview_extraction_completed",
        file_key=file_key,
        chars_extracted=len(preview),
    )
    return preview


def _get_extension(file_key: str) -> str:
    """Extract lowercase file extension from a file key."""
    if "." in file_key:
        return file_key.rsplit(".", 1)[-1].lower()
    return ""


def _extract_markdown(file_bytes: bytes) -> str:
    """Extract text from Markdown, stripping Hugo frontmatter."""
    text = file_bytes.decode("utf-8", errors="replace")

    # Strip Hugo/YAML frontmatter (--- delimited block at start)
    if text.startswith("---"):
        end_idx = text.find("---", 3)
        if end_idx != -1:
            text = text[end_idx + 3:].strip()

    return text


def _extract_text(file_bytes: bytes) -> str:
    """Extract text from plain text files."""
    return file_bytes.decode("utf-8", errors="replace")


def _extract_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF using PyMuPDF (fitz).

    Falls back to OCR for pages with no extractable text (scanned pages).
    Iterates pages until we have enough characters for preview.
    """
    import fitz  # PyMuPDF

    text_parts = []
    total_chars = 0

    with fitz.open(stream=file_bytes, filetype="pdf") as doc:
        for page_num, page in enumerate(doc):
            page_text = page.get_text("text").strip()

            if not page_text:
                # Fallback to OCR for scanned pages
                logger.info(
                    "pdf_page_has_no_text_falling_back_to_ocr",
                    page=page_num + 1,
                )
                pix = page.get_pixmap(dpi=300)
                image_bytes = pix.tobytes("png")
                page_text = _extract_image_ocr(image_bytes).strip()

            if page_text:
                text_parts.append(page_text)
                total_chars += len(page_text)

            if total_chars >= MAX_PREVIEW_CHARS:
                break

    return "\n".join(text_parts)


def _extract_docx(file_bytes: bytes) -> str:
    """Extract text from DOCX using python-docx.

    Iterates paragraphs until we have enough characters for preview.
    """
    from docx import Document

    doc = Document(io.BytesIO(file_bytes))
    text_parts = []
    total_chars = 0

    for para in doc.paragraphs:
        para_text = para.text.strip()
        if para_text:
            text_parts.append(para_text)
            total_chars += len(para_text)
            if total_chars >= MAX_PREVIEW_CHARS:
                break

    return "\n".join(text_parts)


def _extract_xlsx(file_bytes: bytes) -> str:
    """Extract text from XLSX using openpyxl.

    Converts cells to a simple markdown-like table format.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    text_parts = []
    total_chars = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"## Sheet: {sheet_name}\n")

        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            row_text = " | ".join(cells)
            text_parts.append(row_text)
            total_chars += len(row_text)
            if total_chars >= MAX_PREVIEW_CHARS:
                break

        if total_chars >= MAX_PREVIEW_CHARS:
            break

    wb.close()
    return "\n".join(text_parts)


def _extract_image_ocr(file_bytes: bytes) -> str:
    """Extract text from images using pytesseract OCR."""
    from PIL import Image
    import pytesseract
    from src.config.settings import settings

    image = Image.open(io.BytesIO(file_bytes))
    text = pytesseract.image_to_string(image, lang=settings.ocr_language)
    return text
